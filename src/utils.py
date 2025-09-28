#!/usr/bin/env python3
"""
Socratic RAG Enhanced - Core Utilities
=====================================

Comprehensive utility classes for file processing, text analysis, code analysis,
validation, and knowledge extraction. Fully corrected according to project standards.

Classes:
- FileProcessor: Multi-format file processing and content extraction
- TextProcessor: Text analysis, keyword extraction, and processing
- CodeAnalyzer: Code quality analysis and metrics
- ExtendedValidator: Advanced validation utilities
- KnowledgeExtractor: Knowledge extraction from documents
- DocumentParser: Document parsing utilities (backward compatibility)
"""

import os
import re
import ast
import json
import time
import hashlib
import mimetypes
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Tuple
from dataclasses import dataclass, field
from collections import Counter
from enum import Enum

# Core imports
try:
    from src.core import get_logger, get_config, DateTimeHelper, ValidationError, FileHelper

    CORE_AVAILABLE = True
except ImportError:
    CORE_AVAILABLE = False
    import logging
    from datetime import datetime


    def get_logger(name):
        return logging.getLogger(name)


    def get_config():
        return None


    class DateTimeHelper:
        @staticmethod
        def now():
            return datetime.now()

        @staticmethod
        def to_iso_string(dt):
            return dt.isoformat() if dt else None


    class ValidationError(Exception):
        pass


    class FileHelper:
        @staticmethod
        def ensure_directory(path):
            Path(path).mkdir(parents=True, exist_ok=True)

# Optional imports with fallbacks
try:
    from sentence_transformers import SentenceTransformer

    EMBEDDINGS_AVAILABLE = True
except ImportError:
    SentenceTransformer = None
    EMBEDDINGS_AVAILABLE = False

try:
    import PyPDF2

    PDF_AVAILABLE = True
except ImportError:
    PyPDF2 = None
    PDF_AVAILABLE = False

try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    openpyxl = None
    EXCEL_AVAILABLE = False

try:
    import markdown

    MARKDOWN_AVAILABLE = True
except ImportError:
    markdown = None
    MARKDOWN_AVAILABLE = False

try:
    from bs4 import BeautifulSoup

    BS4_AVAILABLE = True
except ImportError:
    BeautifulSoup = None
    BS4_AVAILABLE = False

try:
    import pylint.lint
    from io import StringIO

    PYLINT_AVAILABLE = True
except ImportError:
    pylint = None
    PYLINT_AVAILABLE = False

# Initialize logger
logger = get_logger('utils') if CORE_AVAILABLE else logging.getLogger('utils')


# ============================================================================
# ENUMERATIONS
# ============================================================================

class FileType(Enum):
    """Supported file types"""
    PYTHON = "python"
    JAVASCRIPT = "javascript"
    TYPESCRIPT = "typescript"
    HTML = "html"
    CSS = "css"
    SQL = "sql"
    JSON = "json"
    YAML = "yaml"
    MARKDOWN = "markdown"
    XML = "xml"
    TEXT = "text"
    BINARY = "binary"
    CONFIG = "config"
    DOCKERFILE = "dockerfile"


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class DocumentInfo:
    """Information extracted from a document"""
    file_path: str
    file_name: str
    file_type: str
    mime_type: str
    size_bytes: int
    encoding: Optional[str] = None
    language: Optional[str] = None
    creation_date: Optional[datetime] = None
    modification_date: Optional[datetime] = None

    # Content information
    content: str = ""
    content_hash: str = ""
    page_count: int = 0
    word_count: int = 0
    character_count: int = 0

    # Metadata
    title: Optional[str] = None
    author: Optional[str] = None
    subject: Optional[str] = None
    keywords: List[str] = field(default_factory=list)

    # Processing information
    extraction_method: str = ""
    processing_time: float = 0.0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class TextChunk:
    """A chunk of text with metadata for knowledge processing"""
    chunk_id: str
    content: str
    source_file: str
    source_type: str
    chunk_index: int
    start_position: int
    end_position: int

    # Content analysis
    word_count: int = 0
    character_count: int = 0
    sentence_count: int = 0

    # Context information
    section_title: Optional[str] = None
    page_number: Optional[int] = None
    line_number: Optional[int] = None

    # Processing metadata
    embedding_vector: Optional[List[float]] = None
    keywords: List[str] = field(default_factory=list)
    topics: List[str] = field(default_factory=list)
    importance_score: float = 0.0

    # Timestamps
    created_at: datetime = field(default_factory=DateTimeHelper.now)


@dataclass
class CodeAnalysisResult:
    """Result of code analysis"""
    file_path: str
    file_type: FileType
    analysis_timestamp: datetime = field(default_factory=DateTimeHelper.now)

    # Basic metrics
    lines_of_code: int = 0
    blank_lines: int = 0
    comment_lines: int = 0
    total_lines: int = 0

    # Complexity metrics
    cyclomatic_complexity: float = 0.0
    maintainability_index: float = 0.0
    technical_debt_ratio: float = 0.0

    # Quality indicators
    code_style_score: float = 0.0
    documentation_coverage: float = 0.0
    test_coverage: float = 0.0

    # Issues and suggestions
    syntax_errors: List[Dict[str, Any]] = field(default_factory=list)
    style_issues: List[Dict[str, Any]] = field(default_factory=list)
    security_issues: List[Dict[str, Any]] = field(default_factory=list)
    performance_issues: List[Dict[str, Any]] = field(default_factory=list)
    suggestions: List[str] = field(default_factory=list)

    # Dependencies
    imports: List[str] = field(default_factory=list)
    functions: List[str] = field(default_factory=list)
    classes: List[str] = field(default_factory=list)


# ============================================================================
# FILE PROCESSING UTILITIES
# ============================================================================

class FileProcessor:
    """File processing and content extraction utilities"""

    def __init__(self):
        self.logger = get_logger('utils.file_processor')
        self.config = get_config()

        # Initialize embeddings model if available
        self.embeddings_model = None
        if EMBEDDINGS_AVAILABLE and SentenceTransformer:
            try:
                model_name = getattr(self.config, 'embedding_model',
                                     'all-MiniLM-L6-v2') if self.config else 'all-MiniLM-L6-v2'
                self.embeddings_model = SentenceTransformer(model_name)
                self.logger.info(f"Loaded embeddings model: {model_name}")
            except Exception as e:
                self.logger.warning(f"Failed to load embeddings model: {e}")

    def process_file(self, file_path: str) -> DocumentInfo:
        """Process a file and extract all available information"""
        start_time = DateTimeHelper.now()

        try:
            # Basic file information
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                raise ValidationError(f"File not found: {file_path}")

            # Get file stats
            stat_info = file_path_obj.stat()
            mime_type, _ = mimetypes.guess_type(str(file_path))

            doc_info = DocumentInfo(
                file_path=str(file_path),
                file_name=file_path_obj.name,
                file_type=file_path_obj.suffix.lower(),
                mime_type=mime_type or 'application/octet-stream',
                size_bytes=stat_info.st_size,
                creation_date=datetime.fromtimestamp(stat_info.st_ctime),
                modification_date=datetime.fromtimestamp(stat_info.st_mtime)
            )

            # Extract content based on file type
            self._extract_content(doc_info)

            # Generate content hash
            if doc_info.content:
                doc_info.content_hash = hashlib.sha256(doc_info.content.encode('utf-8')).hexdigest()

            # Calculate processing time
            end_time = DateTimeHelper.now()
            doc_info.processing_time = (end_time - start_time).total_seconds()

            self.logger.info(f"Processed file: {file_path_obj.name} ({doc_info.word_count} words)")
            return doc_info

        except Exception as e:
            self.logger.error(f"Failed to process file {file_path}: {e}")
            raise ValidationError(f"File processing failed: {e}")

    def _extract_content(self, doc_info: DocumentInfo) -> None:
        """Extract content based on file type"""
        file_ext = doc_info.file_type.lower()

        if file_ext == '.txt':
            self._extract_text_content(doc_info)
        elif file_ext == '.md':
            self._extract_markdown_content(doc_info)
        elif file_ext == '.pdf':
            self._extract_pdf_content(doc_info)
        elif file_ext in ['.docx', '.doc']:
            self._extract_docx_content(doc_info)
        elif file_ext in ['.html', '.htm']:
            self._extract_html_content(doc_info)
        elif file_ext in ['.py', '.js', '.ts', '.css', '.sql', '.yaml', '.yml', '.json']:
            self._extract_code_content(doc_info)
        elif file_ext in ['.xlsx', '.xls']:
            self._extract_excel_content(doc_info)
        else:
            # Try to read as text
            try:
                self._extract_text_content(doc_info)
            except Exception:
                doc_info.errors.append(f"Unsupported file type: {file_ext}")

        # Calculate word and character counts
        if doc_info.content:
            doc_info.word_count = len(doc_info.content.split())
            doc_info.character_count = len(doc_info.content)

    def _extract_text_content(self, doc_info: DocumentInfo) -> None:
        """Extract content from plain text files"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']

            for encoding in encodings:
                try:
                    with open(doc_info.file_path, 'r', encoding=encoding) as f:
                        doc_info.content = f.read()
                        doc_info.encoding = encoding
                        doc_info.extraction_method = f"text_reader_{encoding}"
                        break
                except UnicodeDecodeError:
                    continue

            if not doc_info.content:
                raise ValidationError("Could not decode text file with any encoding")

        except Exception as e:
            doc_info.errors.append(f"Text extraction failed: {e}")

    def _extract_markdown_content(self, doc_info: DocumentInfo) -> None:
        """Extract content from Markdown files"""
        try:
            # First extract as text
            self._extract_text_content(doc_info)

            if MARKDOWN_AVAILABLE and markdown and doc_info.content:
                # Convert markdown to HTML for better parsing
                html_content = markdown.markdown(doc_info.content)

                if BS4_AVAILABLE and BeautifulSoup:
                    # Extract plain text from HTML
                    soup = BeautifulSoup(html_content, 'html.parser')
                    doc_info.content = soup.get_text(separator='\n')

                doc_info.extraction_method = "markdown_parser"
            else:
                doc_info.warnings.append("Markdown parsing not available, using text extraction")

        except Exception as e:
            doc_info.errors.append(f"Markdown extraction failed: {e}")

    def _extract_pdf_content(self, doc_info: DocumentInfo) -> None:
        """Extract content from PDF files"""
        if not PDF_AVAILABLE or not PyPDF2:
            doc_info.errors.append("PDF processing not available - PyPDF2 not installed")
            return

        try:
            with open(doc_info.file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)

                # Extract metadata
                if pdf_reader.metadata:
                    doc_info.title = pdf_reader.metadata.get('/Title', '')
                    doc_info.author = pdf_reader.metadata.get('/Author', '')
                    doc_info.subject = pdf_reader.metadata.get('/Subject', '')

                # Extract text from all pages
                doc_info.page_count = len(pdf_reader.pages)
                text_content = []

                for page_num, page in enumerate(pdf_reader.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text.strip():
                            text_content.append(page_text)
                    except Exception as e:
                        doc_info.warnings.append(f"Could not extract text from page {page_num + 1}: {e}")

                doc_info.content = '\n\n'.join(text_content)
                doc_info.extraction_method = "PyPDF2"

        except Exception as e:
            doc_info.errors.append(f"PDF extraction failed: {e}")

    def _extract_docx_content(self, doc_info: DocumentInfo) -> None:
        """Extract content from DOCX files"""
        # Simplified extraction - would need python-docx library
        doc_info.errors.append("DOCX processing not implemented - python-docx library required")

    def _extract_html_content(self, doc_info: DocumentInfo) -> None:
        """Extract content from HTML files"""
        try:
            with open(doc_info.file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()

            if BS4_AVAILABLE and BeautifulSoup:
                soup = BeautifulSoup(html_content, 'html.parser')

                # Extract title
                title_tag = soup.find('title')
                if title_tag:
                    doc_info.title = title_tag.get_text().strip()

                # Extract text content
                doc_info.content = soup.get_text(separator='\n')
                doc_info.extraction_method = "beautifulsoup"
            else:
                # Fallback: basic HTML tag removal
                doc_info.content = re.sub(r'<[^>]+>', '', html_content)
                doc_info.extraction_method = "regex_html_strip"
                doc_info.warnings.append("BeautifulSoup not available, using basic HTML parsing")

        except Exception as e:
            doc_info.errors.append(f"HTML extraction failed: {e}")

    def _extract_code_content(self, doc_info: DocumentInfo) -> None:
        """Extract content from code files"""
        try:
            self._extract_text_content(doc_info)
            doc_info.extraction_method = f"code_reader_{doc_info.file_type}"

        except Exception as e:
            doc_info.errors.append(f"Code extraction failed: {e}")

    def _extract_excel_content(self, doc_info: DocumentInfo) -> None:
        """Extract content from Excel files"""
        if not EXCEL_AVAILABLE or not openpyxl:
            doc_info.errors.append("Excel processing not available - openpyxl not installed")
            return

        try:
            workbook = openpyxl.load_workbook(doc_info.file_path, data_only=True)

            # Extract content from all worksheets
            sheet_contents = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_content = [f"--- Sheet: {sheet_name} ---"]

                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                        sheet_content.append(row_text)

                sheet_contents.append('\n'.join(sheet_content))

            doc_info.content = '\n\n'.join(sheet_contents)
            doc_info.extraction_method = "openpyxl"

        except Exception as e:
            doc_info.errors.append(f"Excel extraction failed: {e}")


# ============================================================================
# TEXT PROCESSING UTILITIES
# ============================================================================

class TextProcessor:
    """Text analysis and processing utilities"""

    def __init__(self):
        self.logger = get_logger('utils.text_processor')
        self.stop_words = {
            'english': {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by',
                        'from', 'this', 'that', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they',
                        'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does',
                        'did', 'will', 'would', 'could', 'should', 'may', 'might', 'must', 'can', 'cannot'}
        }

    def extract_keywords(self, text: str, max_keywords: int = 10, language: str = 'english') -> List[str]:
        """Extract keywords from text using frequency analysis"""
        try:
            if not text.strip():
                return []

            # Clean and tokenize text
            words = self._tokenize_text(text)

            # Remove stop words
            stop_words = self.stop_words.get(language, set())
            filtered_words = [word for word in words if word.lower() not in stop_words and len(word) > 2]

            # Count frequency
            word_freq = Counter(filtered_words)

            # Return top keywords
            return [word for word, _ in word_freq.most_common(max_keywords)]

        except Exception as e:
            self.logger.error(f"Keyword extraction failed: {e}")
            return []

    def extract_sentences(self, text: str) -> List[str]:
        """Extract sentences from text"""
        try:
            # Simple sentence splitting
            sentences = re.split(r'[.!?]+', text)
            return [s.strip() for s in sentences if s.strip()]

        except Exception as e:
            self.logger.error(f"Sentence extraction failed: {e}")
            return []

    def calculate_readability(self, text: str) -> Dict[str, float]:
        """Calculate readability metrics"""
        try:
            if not text.strip():
                return {'flesch_score': 0.0, 'avg_sentence_length': 0.0, 'avg_word_length': 0.0}

            sentences = self.extract_sentences(text)
            words = self._tokenize_text(text)

            if not sentences or not words:
                return {'flesch_score': 0.0, 'avg_sentence_length': 0.0, 'avg_word_length': 0.0}

            # Calculate metrics
            avg_sentence_length = len(words) / len(sentences)
            avg_word_length = sum(len(word) for word in words) / len(words)

            # Simplified Flesch score approximation
            flesch_score = 206.835 - (1.015 * avg_sentence_length) - (84.6 * (avg_word_length / 4.7))
            flesch_score = max(0, min(100, flesch_score))  # Clamp between 0-100

            return {
                'flesch_score': round(flesch_score, 2),
                'avg_sentence_length': round(avg_sentence_length, 2),
                'avg_word_length': round(avg_word_length, 2)
            }

        except Exception as e:
            self.logger.error(f"Readability calculation failed: {e}")
            return {'flesch_score': 0.0, 'avg_sentence_length': 0.0, 'avg_word_length': 0.0}

    def create_chunks(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[TextChunk]:
        """Create overlapping text chunks"""
        try:
            if not text.strip():
                return []

            chunks = []
            text_length = len(text)
            chunk_index = 0
            start = 0

            while start < text_length:
                end = min(start + chunk_size, text_length)
                chunk_content = text[start:end]

                # Try to break at word boundaries
                if end < text_length:
                    last_space = chunk_content.rfind(' ')
                    if last_space > chunk_size * 0.8:  # Don't break too early
                        chunk_content = chunk_content[:last_space]
                        end = start + last_space

                chunk = TextChunk(
                    chunk_id=f"chunk_{chunk_index}",
                    content=chunk_content.strip(),
                    source_file="unknown",
                    source_type="text",
                    chunk_index=chunk_index,
                    start_position=start,
                    end_position=end,
                    word_count=len(chunk_content.split()),
                    character_count=len(chunk_content),
                    sentence_count=len(self.extract_sentences(chunk_content))
                )

                chunks.append(chunk)
                chunk_index += 1
                start = end - overlap if end < text_length else end

            return chunks

        except Exception as e:
            self.logger.error(f"Text chunking failed: {e}")
            return []

    def _tokenize_text(self, text: str) -> List[str]:
        """Simple text tokenization"""
        # Remove punctuation and split into words
        words = re.findall(r'\b[a-zA-Z]+\b', text.lower())
        return [word for word in words if len(word) > 1]


# ============================================================================
# CODE ANALYSIS UTILITIES
# ============================================================================

class CodeAnalyzer:
    """Code analysis and quality assessment utilities"""

    def __init__(self):
        self.logger = get_logger('utils.code_analyzer')
        self.config = get_config()

    def analyze_code_file(self, file_path: str, content: Optional[str] = None) -> CodeAnalysisResult:
        """Analyze a code file and return quality metrics"""
        try:
            # Determine file type
            file_path_obj = Path(file_path)
            file_ext = file_path_obj.suffix.lower()
            file_type = self._get_file_type_from_extension(file_ext)

            # Read content if not provided
            if content is None:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                except Exception as e:
                    raise ValidationError(f"Could not read file {file_path}: {e}")

            result = CodeAnalysisResult(
                file_path=str(file_path),
                file_type=file_type
            )

            # Basic line counting
            self._count_lines(content, result)

            # Language-specific analysis
            if file_ext == '.py':
                self._analyze_python_code(content, result)
            elif file_ext in ['.js', '.ts']:
                self._analyze_javascript_code(content, result)
            elif file_ext == '.sql':
                self._analyze_sql_code(content, result)
            else:
                self._analyze_generic_code(content, result)

            # Common analysis for all file types
            self._analyze_documentation(content, result)
            self._check_code_style(content, result, file_ext)

            self.logger.info(f"Analyzed code file: {file_path_obj.name}")
            return result

        except Exception as e:
            self.logger.error(f"Code analysis failed for {file_path}: {e}")
            raise ValidationError(f"Code analysis failed: {e}")

    def _get_file_type_from_extension(self, ext: str) -> FileType:
        """Map file extension to FileType enum"""
        mapping = {
            '.py': FileType.PYTHON,
            '.js': FileType.JAVASCRIPT,
            '.ts': FileType.TYPESCRIPT,
            '.html': FileType.HTML,
            '.css': FileType.CSS,
            '.sql': FileType.SQL,
            '.json': FileType.JSON,
            '.yaml': FileType.YAML,
            '.yml': FileType.YAML,
            '.md': FileType.MARKDOWN,
            'dockerfile': FileType.DOCKERFILE
        }
        return mapping.get(ext.lower(), FileType.CONFIG)

    def _count_lines(self, content: str, result: CodeAnalysisResult) -> None:
        """Count different types of lines in code"""
        lines = content.split('\n')
        result.total_lines = len(lines)

        for line in lines:
            stripped = line.strip()
            if not stripped:
                result.blank_lines += 1
            elif stripped.startswith('#') or stripped.startswith('//') or stripped.startswith('/*'):
                result.comment_lines += 1
            else:
                result.lines_of_code += 1

    def _analyze_python_code(self, content: str, result: CodeAnalysisResult) -> None:
        """Python-specific code analysis"""
        try:
            # Parse AST
            tree = ast.parse(content)

            # Extract functions and classes
            for node in ast.walk(tree):
                if isinstance(node, ast.FunctionDef):
                    result.functions.append(node.name)
                elif isinstance(node, ast.ClassDef):
                    result.classes.append(node.name)
                elif isinstance(node, ast.Import):
                    for alias in node.names:
                        result.imports.append(alias.name)
                elif isinstance(node, ast.ImportFrom):
                    if node.module:
                        for alias in node.names:
                            result.imports.append(f"{node.module}.{alias.name}")

            # Calculate cyclomatic complexity (simplified)
            complexity = 1  # Base complexity
            for node in ast.walk(tree):
                if isinstance(node, (ast.If, ast.For, ast.While, ast.Try, ast.With, ast.AsyncWith)):
                    complexity += 1
                elif isinstance(node, ast.ExceptHandler):
                    complexity += 1

            result.cyclomatic_complexity = complexity

        except SyntaxError as e:
            result.syntax_errors.append({
                'line': e.lineno,
                'message': str(e),
                'type': 'syntax_error'
            })
        except Exception as e:
            self.logger.warning(f"Python AST analysis failed: {e}")

    def _analyze_javascript_code(self, content: str, result: CodeAnalysisResult) -> None:
        """JavaScript-specific code analysis"""
        # Basic regex-based analysis
        functions = re.findall(r'function\s+(\w+)', content)
        result.functions.extend(functions)

        classes = re.findall(r'class\s+(\w+)', content)
        result.classes.extend(classes)

        imports = re.findall(r'import.*from\s+[\'"]([^\'"]+)[\'"]', content)
        result.imports.extend(imports)

    def _analyze_sql_code(self, content: str, result: CodeAnalysisResult) -> None:
        """SQL-specific code analysis"""
        # Extract table names and operations
        tables = re.findall(r'FROM\s+(\w+)', content, re.IGNORECASE)
        result.functions.extend(tables)  # Store table names in functions list

        operations = re.findall(r'\b(SELECT|INSERT|UPDATE|DELETE|CREATE|DROP|ALTER)\b', content, re.IGNORECASE)
        result.classes.extend(operations)  # Store operations in classes list

    def _analyze_generic_code(self, content: str, result: CodeAnalysisResult) -> None:
        """Generic code analysis for unknown file types"""
        # Basic analysis for any code
        lines = content.split('\n')
        non_empty_lines = [line for line in lines if line.strip()]

        # Estimate complexity based on control flow keywords
        complexity_keywords = ['if', 'else', 'for', 'while', 'switch', 'case', 'try', 'catch', 'except']
        complexity = 1
        for line in non_empty_lines:
            for keyword in complexity_keywords:
                if keyword in line.lower():
                    complexity += 1

        result.cyclomatic_complexity = complexity

    def _analyze_documentation(self, content: str, result: CodeAnalysisResult) -> None:
        """Analyze documentation coverage"""
        total_lines = result.lines_of_code + result.comment_lines
        if total_lines > 0:
            result.documentation_coverage = (result.comment_lines / total_lines) * 100
        else:
            result.documentation_coverage = 0.0

    def _check_code_style(self, content: str, result: CodeAnalysisResult, file_ext: str) -> None:
        """Check code style issues"""
        lines = content.split('\n')

        for i, line in enumerate(lines, 1):
            # Check line length
            if len(line) > 120:
                result.style_issues.append({
                    'line': i,
                    'message': f'Line too long ({len(line)} characters)',
                    'type': 'line_length'
                })

            # Check for trailing whitespace
            if line.rstrip() != line:
                result.style_issues.append({
                    'line': i,
                    'message': 'Trailing whitespace',
                    'type': 'whitespace'
                })


# ============================================================================
# EXTENDED VALIDATION UTILITIES
# ============================================================================

class ExtendedValidator:
    """Extended validation utilities beyond core ValidationHelper"""

    def __init__(self):
        self.logger = get_logger('utils.validator')

    def validate_project_specifications(self, project_data: Dict[str, Any]) -> List[str]:
        """Validate project specifications for completeness"""
        issues = []

        # Basic validation
        name = project_data.get('name', '').strip()
        if not name:
            issues.append("Project name is required")

        owner_id = project_data.get('owner_id', '').strip()
        if not owner_id:
            issues.append("Project owner is required")

        # Technology stack validation
        tech_stack = project_data.get('technology_stack', {})
        if tech_stack and not isinstance(tech_stack, dict):
            issues.append("Technology stack must be a dictionary")

        for tech in tech_stack.values():
            if not self.validate_technology_name(str(tech)):
                issues.append(f"Technology '{tech}' format not recognized")

        return issues

    def validate_technology_name(self, tech: str) -> bool:
        """Validate technology name format"""
        if not tech or not tech.strip():
            return False

        # Allow alphanumeric, dots, hyphens, and spaces
        pattern = r'^[a-zA-Z0-9\s\.\-_+]+$'
        return bool(re.match(pattern, tech.strip()))

    def validate_file_path(self, file_path: str) -> bool:
        """Validate file path format"""
        if not file_path:
            return False

        # Check for invalid characters
        invalid_chars = ['<', '>', ':', '"', '|', '?', '*']
        return not any(char in file_path for char in invalid_chars)

    def validate_code_content(self, content: str, file_type: FileType) -> List[str]:
        """Validate code content"""
        issues = []

        if not content.strip():
            issues.append("Code content is empty")
            return issues

        # Basic syntax validation
        if file_type == FileType.PYTHON:
            try:
                ast.parse(content)
            except SyntaxError as e:
                issues.append(f"Python syntax error: {e}")

        elif file_type == FileType.JSON:
            try:
                json.loads(content)
            except json.JSONDecodeError as e:
                issues.append(f"JSON syntax error: {e}")

        return issues


# ============================================================================
# KNOWLEDGE EXTRACTION UTILITIES
# ============================================================================

class KnowledgeExtractor:
    """Knowledge extraction from documents and text"""

    def __init__(self):
        self.logger = get_logger('utils.knowledge_extractor')

    def extract_entities(self, text: str) -> Dict[str, List[str]]:
        """Extract entities from text using pattern matching"""
        try:
            entities = {
                'emails': [],
                'urls': [],
                'phone_numbers': [],
                'dates': [],
                'currencies': []
            }

            # Email extraction
            email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            entities['emails'] = re.findall(email_pattern, text)

            # URL extraction
            url_pattern = r'https?://[^\s]+'
            entities['urls'] = re.findall(url_pattern, text)

            # Phone number extraction (simple pattern)
            phone_pattern = r'\b\d{3}-\d{3}-\d{4}\b|\b\(\d{3}\)\s*\d{3}-\d{4}\b'
            entities['phone_numbers'] = re.findall(phone_pattern, text)

            # Date extraction (simple patterns)
            date_pattern = r'\b\d{1,2}/\d{1,2}/\d{4}\b|\b\d{4}-\d{2}-\d{2}\b'
            entities['dates'] = re.findall(date_pattern, text)

            # Currency extraction
            currency_pattern = r'\$\d+(?:\.\d{2})?|\$\d{1,3}(?:,\d{3})*(?:\.\d{2})?'
            entities['currencies'] = re.findall(currency_pattern, text)

            return entities

        except Exception as e:
            self.logger.error(f"Entity extraction failed: {e}")
            return {'emails': [], 'urls': [], 'phone_numbers': [], 'dates': [], 'currencies': []}

    def extract_concepts(self, text: str) -> List[str]:
        """Extract key concepts from text"""
        try:
            # Simple concept extraction using noun phrases
            words = re.findall(r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\b', text)

            # Filter and deduplicate
            concepts = list(set(words))
            return concepts[:20]  # Limit to top 20

        except Exception as e:
            self.logger.error(f"Concept extraction failed: {e}")
            return []

    def extract_requirements(self, text: str) -> List[str]:
        """Extract requirements from text"""
        try:
            requirements = []

            # Pattern for requirement statements
            req_patterns = [
                r'(?:must|shall|should|will)\s+([^.!?]+)',
                r'(?:requirement|need|necessary)\s*:\s*([^.!?]+)',
                r'(?:the system|application|software)\s+(?:must|shall|should|will)\s+([^.!?]+)'
            ]

            for pattern in req_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    clean_req = match.strip()
                    if len(clean_req) > 10:
                        requirements.append(clean_req)

            return requirements

        except Exception as e:
            self.logger.error(f"Requirements extraction failed: {e}")
            return []

    def extract_risks(self, text: str) -> List[str]:
        """Extract risks from text"""
        try:
            risks = []

            # Pattern for risk statements
            risk_patterns = [
                r'(?:risk|danger|problem|issue)\s*:\s*([^.!?]+)',
                r'(?:might|could|may)\s+(?:fail|break|cause problems?)\s+([^.!?]+)',
                r'(?:potential|possible)\s+(?:issue|problem|risk)\s+([^.!?]+)',
            ]

            for pattern in risk_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    clean_risk = match.strip()
                    if len(clean_risk) > 10:
                        risks.append(clean_risk)

            return risks

        except Exception as e:
            self.logger.error(f"Risk extraction failed: {e}")
            return []


# ============================================================================
# DOCUMENT PARSER (BACKWARD COMPATIBILITY)
# ============================================================================

class DocumentParser:
    """Document parsing utilities (alias for FileProcessor for backward compatibility)"""

    def __init__(self):
        self.logger = get_logger('utils.document_parser')
        self.file_processor = FileProcessor()

    def parse_document(self, file_path: str) -> DocumentInfo:
        """Parse a document and return extracted information"""
        return self.file_processor.process_file(file_path)


# ============================================================================
# UTILITY FACTORY AND MAIN INTERFACE
# ============================================================================

class UtilityFactory:
    """Factory for creating utility instances"""

    _instances = {}

    @classmethod
    def get_file_processor(cls) -> FileProcessor:
        """Get FileProcessor instance (singleton)"""
        if 'file_processor' not in cls._instances:
            cls._instances['file_processor'] = FileProcessor()
        return cls._instances['file_processor']

    @classmethod
    def get_text_processor(cls) -> TextProcessor:
        """Get TextProcessor instance (singleton)"""
        if 'text_processor' not in cls._instances:
            cls._instances['text_processor'] = TextProcessor()
        return cls._instances['text_processor']

    @classmethod
    def get_code_analyzer(cls) -> CodeAnalyzer:
        """Get CodeAnalyzer instance (singleton)"""
        if 'code_analyzer' not in cls._instances:
            cls._instances['code_analyzer'] = CodeAnalyzer()
        return cls._instances['code_analyzer']

    @classmethod
    def get_validator(cls) -> ExtendedValidator:
        """Get ExtendedValidator instance (singleton)"""
        if 'validator' not in cls._instances:
            cls._instances['validator'] = ExtendedValidator()
        return cls._instances['validator']

    @classmethod
    def get_knowledge_extractor(cls) -> KnowledgeExtractor:
        """Get KnowledgeExtractor instance (singleton)"""
        if 'knowledge_extractor' not in cls._instances:
            cls._instances['knowledge_extractor'] = KnowledgeExtractor()
        return cls._instances['knowledge_extractor']


# ============================================================================
# MAIN EXPORTS
# ============================================================================

def get_file_processor() -> FileProcessor:
    """Get file processor instance"""
    return UtilityFactory.get_file_processor()


def get_text_processor() -> TextProcessor:
    """Get text processor instance"""
    return UtilityFactory.get_text_processor()


def get_code_analyzer() -> CodeAnalyzer:
    """Get code analyzer instance"""
    return UtilityFactory.get_code_analyzer()


def get_validator() -> ExtendedValidator:
    """Get extended validator instance"""
    return UtilityFactory.get_validator()


def get_knowledge_extractor() -> KnowledgeExtractor:
    """Get knowledge extractor instance"""
    return UtilityFactory.get_knowledge_extractor()


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_project_data(project_data: Dict[str, Any]) -> List[str]:
    """Validate project data dictionary - standalone validation"""
    try:
        issues = []

        # Required fields
        required_fields = ['name', 'owner_id']
        for field in required_fields:
            if field not in project_data or not str(project_data[field]).strip():
                issues.append(f"Required field '{field}' is missing or empty")

        # Name validation
        if 'name' in project_data:
            name = str(project_data['name']).strip()
            if len(name) < 2:
                issues.append("Project name must be at least 2 characters long")

        # Status validation
        if 'status' in project_data:
            status = project_data['status']
            valid_statuses = ['active', 'inactive', 'completed', 'archived', 'cancelled']
            if status not in valid_statuses:
                issues.append(f"Invalid status: '{status}'. Must be one of: {valid_statuses}")

        # Phase validation
        if 'phase' in project_data:
            phase = project_data['phase']
            valid_phases = ['planning', 'requirements', 'design', 'development', 'testing', 'deployment', 'maintenance',
                            'completed', 'cancelled']
            if phase not in valid_phases:
                issues.append(f"Invalid phase: '{phase}'. Must be one of: {valid_phases}")

        return issues

    except Exception as e:
        logger.error(f"Project data validation failed: {e}")
        return [f"Validation error: {str(e)}"]


# ============================================================================
# MODULE EXPORTS
# ============================================================================

__all__ = [
    # Data structures
    'DocumentInfo', 'TextChunk', 'CodeAnalysisResult', 'FileType',

    # Main utility classes
    'FileProcessor', 'DocumentParser', 'TextProcessor', 'CodeAnalyzer',
    'ExtendedValidator', 'KnowledgeExtractor',

    # Factory and convenience functions
    'UtilityFactory',
    'get_file_processor', 'get_text_processor', 'get_code_analyzer',
    'get_validator', 'get_knowledge_extractor',

    # Validation functions
    'validate_project_data'
]

# ============================================================================
# MODULE INITIALIZATION AND TESTING
# ============================================================================

if __name__ == "__main__":
    # Test the utilities
    logger.info("Testing utilities module...")

    try:
        # Test file processor
        file_proc = get_file_processor()
        logger.info("✅ FileProcessor created")

        # Test text processor
        text_proc = get_text_processor()
        test_text = "This is a test document with important information about Python development."
        keywords = text_proc.extract_keywords(test_text)
        logger.info(f"✅ TextProcessor test - Keywords: {keywords}")

        # Test code analyzer
        code_analyzer = get_code_analyzer()
        logger.info("✅ CodeAnalyzer created")

        # Test validator
        validator = get_validator()
        test_tech = "Python 3.9"
        is_valid = validator.validate_technology_name(test_tech)
        logger.info(f"✅ Validator test - '{test_tech}' is valid: {is_valid}")

        # Test knowledge extractor
        knowledge = get_knowledge_extractor()
        logger.info("✅ KnowledgeExtractor created")

        # Test validate_project_data function
        test_project_data = {
            'name': 'Test Project',
            'owner_id': 'user123',
            'status': 'active'
        }
        validation_issues = validate_project_data(test_project_data)
        logger.info(f"✅ validate_project_data test - Issues: {validation_issues}")

        logger.info("🎉 All utilities tests passed!")

    except Exception as e:
        logger.error(f"❌ Utilities test failed: {e}")
        raise
